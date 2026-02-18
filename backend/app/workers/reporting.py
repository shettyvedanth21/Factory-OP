"""Report generators for PDF and Excel output."""
import json
import io
from datetime import datetime
from typing import Dict, Any, List, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, ListFlowable, ListItem
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.core.logging import get_logger

logger = get_logger(__name__)


def _create_table_style() -> TableStyle:
    """Create consistent table styling."""
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ])


def generate_pdf(
    report_config: Dict[str, Any],
    data: Dict[str, Any],
    analytics_results: Optional[Dict[str, Any]] = None,
) -> bytes:
    """Generate PDF report.
    
    Args:
        report_config: Report configuration (title, date range, etc.)
        data: Aggregated report data
        analytics_results: Optional analytics job results
        
    Returns:
        PDF bytes
    """
    logger.info("pdf.generating", title=report_config.get("title"))
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=18,
    )
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Page 1: Cover
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#1e40af"),
        spaceAfter=30,
        alignment=1,  # Center
    )
    
    elements.append(Paragraph("Factory Operations Report", title_style))
    elements.append(Spacer(1, 0.3 * inch))
    
    if report_config.get("title"):
        elements.append(Paragraph(report_config["title"], styles["Heading2"]))
    
    metadata = data.get("report_metadata", {})
    cover_info = [
        ["Generated:", metadata.get("generated_at", "N/A")[:19].replace("T", " ")],
        ["Period:", f"{metadata.get('date_range_start', 'N/A')[:10]} to {metadata.get('date_range_end', 'N/A')[:10]}"],
        ["Devices:", str(len(data.get("devices", [])))],
    ]
    
    cover_table = Table(cover_info, colWidths=[2 * inch, 4 * inch])
    cover_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    elements.append(cover_table)
    elements.append(PageBreak())
    
    # Page 2: Executive Summary
    elements.append(Paragraph("Executive Summary", styles["Heading2"]))
    elements.append(Spacer(1, 0.2 * inch))
    
    alert_summary = data.get("alert_summary", {})
    summary_data = [
        ["Metric", "Value"],
        ["Total Devices", str(len(data.get("devices", [])))],
        ["Total Alerts", str(alert_summary.get("total", 0))],
        ["Critical Alerts", str(alert_summary.get("critical", 0))],
        ["High Alerts", str(alert_summary.get("high", 0))],
        ["Medium Alerts", str(alert_summary.get("medium", 0))],
        ["Low Alerts", str(alert_summary.get("low", 0))],
    ]
    
    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(_create_table_style())
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Energy overview (if power data available)
    telemetry_summary = data.get("telemetry_summary", {})
    power_devices = []
    for device_key, params in telemetry_summary.items():
        if "power" in params:
            power_devices.append([
                device_key.replace("device_", "Device "),
                f"{params['power']['avg']} W (avg)",
                f"{params['power']['min']} - {params['power']['max']} W",
            ])
    
    if power_devices:
        elements.append(Paragraph("Energy Overview", styles["Heading3"]))
        elements.append(Spacer(1, 0.1 * inch))
        power_table = Table(
            [["Device", "Average Power", "Power Range"]] + power_devices,
            colWidths=[2 * inch, 2 * inch, 2 * inch],
        )
        power_table.setStyle(_create_table_style())
        elements.append(power_table)
    
    elements.append(PageBreak())
    
    # Page 3: Device Details
    elements.append(Paragraph("Device Telemetry Summary", styles["Heading2"]))
    elements.append(Spacer(1, 0.2 * inch))
    
    for device in data.get("devices", []):
        device_id = device["id"]
        device_key = f"device_{device_id}"
        
        elements.append(Paragraph(
            f"{device.get('name', device['device_key'])} ({device['device_key']})",
            styles["Heading3"]
        ))
        
        device_stats = telemetry_summary.get(device_key, {})
        if device_stats:
            stats_data = [["Parameter", "Min", "Max", "Average", "Samples"]]
            for param, stats in device_stats.items():
                stats_data.append([
                    param,
                    str(stats.get("min", "N/A")),
                    str(stats.get("max", "N/A")),
                    str(stats.get("avg", "N/A")),
                    str(stats.get("count", 0)),
                ])
            
            stats_table = Table(stats_data, colWidths=[1.5 * inch, 1 * inch, 1 * inch, 1 * inch, 1 * inch])
            stats_table.setStyle(_create_table_style())
            elements.append(stats_table)
        else:
            elements.append(Paragraph("No telemetry data available for this period.", styles["Normal"]))
        
        elements.append(Spacer(1, 0.2 * inch))
    
    elements.append(PageBreak())
    
    # Page 4: Alerts Log
    elements.append(Paragraph("Alerts Log", styles["Heading2"]))
    elements.append(Spacer(1, 0.2 * inch))
    
    alerts = data.get("alerts", [])
    if alerts:
        # Sort by triggered_at descending
        sorted_alerts = sorted(alerts, key=lambda x: x.get("triggered_at", ""), reverse=True)
        
        alerts_data = [["Time", "Severity", "Device ID", "Message"]]
        for alert in sorted_alerts[:100]:  # Limit to 100 alerts
            alerts_data.append([
                alert.get("triggered_at", "N/A")[:16].replace("T", " "),
                alert.get("severity", "N/A").upper(),
                str(alert.get("device_id", "N/A")),
                alert.get("message", "N/A")[:60],
            ])
        
        alerts_table = Table(alerts_data, colWidths=[1.2 * inch, 0.9 * inch, 0.9 * inch, 3 * inch])
        alerts_table.setStyle(_create_table_style())
        elements.append(alerts_table)
    else:
        elements.append(Paragraph("No alerts recorded during this period.", styles["Normal"]))
    
    # Page 5+: Analytics Results
    if analytics_results and report_config.get("include_analytics"):
        elements.append(PageBreak())
        elements.append(Paragraph("Analytics Results", styles["Heading2"]))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Summary text
        if "summary" in analytics_results:
            elements.append(Paragraph(analytics_results["summary"], styles["Normal"]))
            elements.append(Spacer(1, 0.2 * inch))
        
        # Anomaly results
        if "anomalies" in analytics_results:
            elements.append(Paragraph("Detected Anomalies", styles["Heading3"]))
            anomalies = analytics_results["anomalies"][:20]  # Limit to 20
            
            if anomalies:
                anomaly_data = [["Timestamp", "Device", "Score"]]
                for anomaly in anomalies:
                    anomaly_data.append([
                        anomaly.get("timestamp", "N/A")[:16].replace("T", " "),
                        str(anomaly.get("device_id", "N/A")),
                        f"{anomaly.get('score', 0):.3f}",
                    ])
                
                anomaly_table = Table(anomaly_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch])
                anomaly_table.setStyle(_create_table_style())
                elements.append(anomaly_table)
            else:
                elements.append(Paragraph("No anomalies detected.", styles["Normal"]))
        
        # Forecast summary
        if "forecast" in analytics_results:
            elements.append(Paragraph("Energy Forecast", styles["Heading3"]))
            forecast = analytics_results.get("forecast", {})
            horizon = forecast.get("horizon_days", "N/A")
            elements.append(Paragraph(f"Forecast horizon: {horizon} days", styles["Normal"]))
    
    # Build PDF
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    logger.info("pdf.generated", size_bytes=len(pdf_bytes))
    return pdf_bytes


def generate_excel(
    report_config: Dict[str, Any],
    data: Dict[str, Any],
    analytics_results: Optional[Dict[str, Any]] = None,
) -> bytes:
    """Generate Excel report.
    
    Args:
        report_config: Report configuration
        data: Aggregated report data
        analytics_results: Optional analytics job results
        
    Returns:
        Excel bytes
    """
    logger.info("excel.generating", title=report_config.get("title"))
    
    wb = Workbook()
    
    # Remove default sheet and create custom ones
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Factory Operations Report"])
    ws_summary.append([])
    ws_summary.append(["Report Title", report_config.get("title", "N/A")])
    ws_summary.append(["Generated", data.get("report_metadata", {}).get("generated_at", "N/A")])
    ws_summary.append(["Period", f"{data.get('report_metadata', {}).get('date_range_start', 'N/A')[:10]} to {data.get('report_metadata', {}).get('date_range_end', 'N/A')[:10]}"])
    ws_summary.append([])
    ws_summary.append(["Key Metrics"])
    alert_summary = data.get("alert_summary", {})
    ws_summary.append(["Total Devices", len(data.get("devices", []))])
    ws_summary.append(["Total Alerts", alert_summary.get("total", 0)])
    ws_summary.append(["Critical", alert_summary.get("critical", 0)])
    ws_summary.append(["High", alert_summary.get("high", 0)])
    ws_summary.append(["Medium", alert_summary.get("medium", 0)])
    ws_summary.append(["Low", alert_summary.get("low", 0)])
    
    # Sheet 2: Devices
    ws_devices = wb.create_sheet("Devices")
    ws_devices.append(["ID", "Device Key", "Name", "Manufacturer", "Model", "Region", "Last Seen"])
    
    # Style header
    for cell in ws_devices[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    for device in data.get("devices", []):
        ws_devices.append([
            device.get("id"),
            device.get("device_key"),
            device.get("name"),
            device.get("manufacturer"),
            device.get("model"),
            device.get("region"),
            device.get("last_seen", "N/A"),
        ])
    
    # Auto-adjust column widths
    for column in ws_devices.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_devices.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3: Alerts
    ws_alerts = wb.create_sheet("Alerts")
    ws_alerts.append(["ID", "Rule ID", "Device ID", "Triggered At", "Severity", "Message"])
    
    # Style header
    for cell in ws_alerts[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    for alert in sorted(data.get("alerts", []), key=lambda x: x.get("triggered_at", ""), reverse=True):
        ws_alerts.append([
            alert.get("id"),
            alert.get("rule_id"),
            alert.get("device_id"),
            alert.get("triggered_at"),
            alert.get("severity"),
            alert.get("message"),
        ])
    
    # Sheet 4: Telemetry Summary
    ws_telemetry = wb.create_sheet("Telemetry")
    ws_telemetry.append(["Device ID", "Parameter", "Min", "Max", "Average", "Samples"])
    
    # Style header
    for cell in ws_telemetry[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    for device_key, params in data.get("telemetry_summary", {}).items():
        device_id = device_key.replace("device_", "")
        for param, stats in params.items():
            ws_telemetry.append([
                device_id,
                param,
                stats.get("min"),
                stats.get("max"),
                stats.get("avg"),
                stats.get("count"),
            ])
    
    # Sheet 5: Analytics (if applicable)
    if analytics_results and report_config.get("include_analytics"):
        ws_analytics = wb.create_sheet("Analytics")
        ws_analytics.append(["Analytics Results"])
        ws_analytics.append([])
        
        if "summary" in analytics_results:
            ws_analytics.append(["Summary", analytics_results["summary"]])
        
        if "anomaly_count" in analytics_results:
            ws_analytics.append([])
            ws_analytics.append(["Anomaly Detection"])
            ws_analytics.append(["Total Anomalies", analytics_results.get("anomaly_count", 0)])
            ws_analytics.append(["Anomaly Score", analytics_results.get("anomaly_score", 0)])
            
            if "anomalies" in analytics_results:
                ws_analytics.append([])
                ws_analytics.append(["Timestamp", "Device ID", "Score"])
                for anomaly in analytics_results["anomalies"][:50]:
                    ws_analytics.append([
                        anomaly.get("timestamp"),
                        anomaly.get("device_id"),
                        anomaly.get("score"),
                    ])
        
        if "forecast" in analytics_results:
            ws_analytics.append([])
            ws_analytics.append(["Energy Forecast"])
            ws_analytics.append(["Horizon (days)", analytics_results.get("horizon_days", "N/A")])
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    logger.info("excel.generated", size_bytes=len(excel_bytes))
    return excel_bytes
